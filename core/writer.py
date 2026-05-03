"""代理店ごとの.xlsx ファイル末尾に新シートを追記する。

openpyxl でクロスプラットフォーム対応（Windows/Linux共通）。
.xls 形式は事前に .xlsx に変換しておくこと（scripts/convert_xls_to_xlsx.py）。
"""
from __future__ import annotations
import os
import shutil
from typing import Dict, List, Callable, Optional
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import CATEGORIES, LIMITED_AGENT_COLUMNS
from .aggregate import UNASSIGNED_LABEL


def _columns_for_agent(agent: str) -> List[str]:
    return LIMITED_AGENT_COLUMNS.get(agent, CATEGORIES)


def find_agent_file(margin_dir: str, agent: str) -> Optional[str]:
    """代理店名に一致する xlsx (なければ xls) ファイルを探す"""
    if not os.path.isdir(margin_dir):
        return None
    target_xlsx = f"カルチャーキッズマージン精算書（{agent}）.xlsx"
    direct = os.path.join(margin_dir, target_xlsx)
    if os.path.exists(direct):
        return direct
    target_xls = f"カルチャーキッズマージン精算書（{agent}）.xls"
    direct_xls = os.path.join(margin_dir, target_xls)
    if os.path.exists(direct_xls):
        return direct_xls
    # 部分一致探索（agentが含まれる、かつbakでない）
    for fn in os.listdir(margin_dir):
        if "_bak_" in fn:
            continue
        if not fn.endswith((".xlsx", ".xls")):
            continue
        if agent in fn:
            return os.path.join(margin_dir, fn)
    return None


def backup_file(path: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(path)
    bak = f"{base}_bak_{ts}{ext}"
    shutil.copy2(path, bak)
    return bak


def _write_sheet(wb: openpyxl.Workbook, agent: str, records: List[Dict],
                 sheet_name: str) -> tuple[str, int]:
    """既存workbookに代理店データシートを追加する。返り値は (実際のシート名, 行数)"""
    cols = _columns_for_agent(agent)
    header = ["家族ID", "塾名", "代理店", "対象月", "入金日", *cols, "合計"]

    # 重複シート名の場合は連番付与
    existing = set(wb.sheetnames)
    target_name = sheet_name
    n = 1
    while target_name in existing:
        n += 1
        target_name = f"{sheet_name}_{n}"
    ws = wb.create_sheet(target_name)

    # ヘッダ
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="FFE699")
    for col_idx, h in enumerate(header, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = bold
        c.fill = fill

    # データ
    total_sum = 0
    for row_offset, r in enumerate(records, start=2):
        row_total = sum(r.get(c, 0) for c in cols)
        ws.cell(row=row_offset, column=1, value=r["家族ID"])
        ws.cell(row=row_offset, column=2, value=r["塾名"])
        ws.cell(row=row_offset, column=3, value=r["代理店"])
        ws.cell(row=row_offset, column=4, value=r["対象月"])
        date_cell = ws.cell(row=row_offset, column=5, value=r["入金日"])
        date_cell.number_format = "yyyy/m/d"
        for ci, cat in enumerate(cols, start=6):
            ws.cell(row=row_offset, column=ci, value=r.get(cat, 0))
        ws.cell(row=row_offset, column=5 + len(cols) + 1, value=row_total)
        total_sum += row_total

    # 売上合計行
    last_row = ws.max_row + 2
    c = ws.cell(row=last_row, column=5, value="売上合計")
    c.font = bold
    c2 = ws.cell(row=last_row, column=5 + len(cols) + 1, value=total_sum)
    c2.font = bold

    # 列幅
    widths = [10, 28, 18, 12, 12] + [16] * len(cols) + [12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    return target_name, len(records)


def write_to_xlsx(margin_dir: str, by_agent: Dict[str, List[Dict]], sheet_name: str,
                  backup: bool = True, create_missing: bool = True,
                  progress: Optional[Callable[[int, int, str], None]] = None) -> Dict[str, str]:
    """openpyxlで各代理店ファイルに新シートを追加（クロスプラットフォーム対応）

    Returns: {代理店名: 結果文字列}
    """
    results: Dict[str, str] = {}
    total = len(by_agent)

    for idx, (agent, records) in enumerate(by_agent.items(), start=1):
        if agent == UNASSIGNED_LABEL:
            results[agent] = "skip:未マッピングのため除外"
            continue
        if progress:
            progress(idx, total, agent)

        path = find_agent_file(margin_dir, agent)
        created = False
        if path is None:
            if not create_missing:
                results[agent] = "skip:ファイルなし"
                continue
            path = os.path.join(margin_dir, f"カルチャーキッズマージン精算書（{agent}）.xlsx")
            wb = openpyxl.Workbook()
            # デフォルトシートを削除（_write_sheetが新規作成する）
            default = wb.active
            wb.remove(default)
            created = True
        else:
            # .xls ファイルを処理する場合は openpyxl では読めないため警告
            if path.endswith(".xls"):
                results[agent] = f"error:.xls形式は非対応 ({os.path.basename(path)})"
                continue
            if backup:
                backup_file(path)
            wb = openpyxl.load_workbook(path)

        try:
            actual_sheet, n_rows = _write_sheet(wb, agent, records, sheet_name)
            # 保存先は常に .xlsx で
            if created:
                wb.save(path)
                results[agent] = f"created:{os.path.basename(path)}"
            else:
                if path.endswith(".xls"):
                    new_path = path[:-4] + ".xlsx"
                    wb.save(new_path)
                    results[agent] = f"updated→xlsx:{os.path.basename(new_path)} sheet '{actual_sheet}' ({n_rows}行)"
                else:
                    wb.save(path)
                    results[agent] = f"updated:{os.path.basename(path)} sheet '{actual_sheet}' ({n_rows}行)"
        finally:
            wb.close()

    return results


# 後方互換のエイリアス（旧コードがwrite_via_excelを呼んでいる場合）
def write_via_excel(*args, **kwargs):
    """互換ラッパー: 旧 pywin32 ベースの関数名を openpyxl 実装に流す"""
    return write_to_xlsx(*args, **kwargs)
