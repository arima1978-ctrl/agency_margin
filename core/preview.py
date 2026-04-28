"""集計結果を1つのプレビューxlsxに書き出す（代理店ごとシート＋対応表＋未マッピング）"""
from __future__ import annotations
import os
from typing import Dict, List

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import CATEGORIES, LIMITED_AGENT_COLUMNS
from .aggregate import UNASSIGNED_LABEL, agent_totals


def _columns_for_agent(agent: str) -> List[str]:
    return LIMITED_AGENT_COLUMNS.get(agent, CATEGORIES)


def _safe_sheet_name(name: str) -> str:
    bad = '[]:*?/\\'
    s = "".join("_" if c in bad else c for c in name)
    return s[:31] or "(空)"


def write_preview(out_path: str, by_agent: Dict[str, List[Dict]], quarter_label: str) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="FFE699")

    # 対応表シート（先頭）— 未設定は除外
    ws_idx = wb.create_sheet("_対応表")
    ws_idx.append(["代理店", "件数", "売上合計"])
    for c in ws_idx[1]:
        c.font = bold
        c.fill = fill
    for agent, n, total in agent_totals(by_agent):
        if agent == UNASSIGNED_LABEL:
            continue
        ws_idx.append([agent, n, total])
    ws_idx.column_dimensions["A"].width = 30
    ws_idx.column_dimensions["B"].width = 10
    ws_idx.column_dimensions["C"].width = 14

    # 代理店シート — 未設定は除外
    sorted_agents = [a for a, _, _ in agent_totals(by_agent) if a != UNASSIGNED_LABEL]
    used = set()
    for agent in sorted_agents:
        cols = _columns_for_agent(agent)
        header = ["家族ID", "塾名", "代理店", "対象月", "入金日", *cols, "合計"]
        sname = _safe_sheet_name(agent)
        # 重複名対策
        base, n = sname, 1
        while sname in used:
            n += 1
            sname = f"{base[:28]}_{n}"
        used.add(sname)
        ws = wb.create_sheet(sname)
        ws.append(header)
        for c in ws[1]:
            c.font = bold
            c.fill = fill
        total_sum = 0
        for r in by_agent[agent]:
            row_total = sum(r.get(c, 0) for c in cols)
            ws.append([
                r["家族ID"], r["塾名"], r["代理店"], r["対象月"], r["入金日"],
                *(r.get(c, 0) for c in cols),
                row_total,
            ])
            total_sum += row_total
        ws.append([])
        last = ["", "", "", "", "売上合計", *("" for _ in cols), total_sum]
        ws.append(last)
        ws.cell(row=ws.max_row, column=5).font = bold
        ws.cell(row=ws.max_row, column=len(header)).font = bold

        widths = [10, 28, 18, 12, 12] + [16] * len(cols) + [12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # 未マッピング詳細
    if UNASSIGNED_LABEL in by_agent:
        ws = wb.create_sheet("_未マッピング詳細")
        ws.append(["家族ID", "塾名", "対象月", "合計"])
        for c in ws[1]:
            c.font = bold
            c.fill = fill
        for r in by_agent[UNASSIGNED_LABEL]:
            ws.append([r["家族ID"], r["塾名"], r["対象月"], r["合計"]])
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path
