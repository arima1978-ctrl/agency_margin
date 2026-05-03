"""送信分xlsmから売上行を抽出する"""
from __future__ import annotations
import os
import warnings
from typing import List, Dict, Optional
from datetime import datetime, date

warnings.filterwarnings("ignore")
import openpyxl

from .config import (
    CATEGORIES,
    COL_KAZOKU_ID,
    COL_RYOKIN,
    COL_JUKUMEI,
    NYUKIN_SHEET,
    NYUKIN_DATA_START,
    NYUKIN_COL_KAZOKU_ID,
    NYUKIN_COL_NYUKIN_DATE,
)


def _to_int(v):
    if v is None or v == "":
        return None
    try:
        if isinstance(v, str):
            v = v.replace(",", "").strip()
            if v == "":
                return None
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _to_money(v) -> int:
    if v is None or v == "":
        return 0
    try:
        if isinstance(v, str):
            v = v.replace(",", "").strip()
            if v == "":
                return 0
        return int(round(float(v)))
    except (TypeError, ValueError):
        return 0


def _coerce_date(v) -> Optional[datetime]:
    """セル値を datetime に変換。日付として認識できなければ None。"""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    # 文字列日付（"2026-02-06" 等）に対応
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def load_paid_map(wb) -> Dict[int, datetime]:
    """⑮入金チェックシートの P列から {家族ID: 入金日} を作る（P列空欄は未入金で除外）"""
    if NYUKIN_SHEET not in wb.sheetnames:
        return {}
    ws = wb[NYUKIN_SHEET]
    paid: Dict[int, datetime] = {}
    for row in ws.iter_rows(min_row=NYUKIN_DATA_START, values_only=True):
        if row is None or len(row) <= max(NYUKIN_COL_KAZOKU_ID, NYUKIN_COL_NYUKIN_DATE):
            continue
        kid = _to_int(row[NYUKIN_COL_KAZOKU_ID])
        if not kid or kid <= 0:
            continue
        d = _coerce_date(row[NYUKIN_COL_NYUKIN_DATE])
        if d is None:
            continue  # 未入金 → 対象外
        # 同じ家族IDが複数行ある場合は最も後の入金日を採用
        if kid not in paid or d > paid[kid]:
            paid[kid] = d
    return paid


def extract_sales(xlsm_path: str, target_month: str) -> List[Dict]:
    """1つの送信分xlsmから 家族ID ごとに集計したレコード配列を返す。

    同一家族IDが複数の塾名で登録されている場合は1行に集約する（金額は合算、
    塾名は最長の表記を採用）。
    入金日は ⑮入金チェックシート P列から取得し、未入金（P列空欄）の家族IDは除外する。
    """
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)
    paid_map = load_paid_map(wb)

    # kid -> {"juku_candidates": set, category amounts...}
    rows: Dict[int, Dict] = {}
    available = set(wb.sheetnames)
    for cat in CATEGORIES:
        if cat not in available:
            continue
        ws = wb[cat]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row is None:
                continue
            if len(row) <= max(COL_KAZOKU_ID, COL_RYOKIN, COL_JUKUMEI):
                continue
            kid = _to_int(row[COL_KAZOKU_ID])
            if not kid or kid <= 0:
                continue
            if kid not in paid_map:
                continue  # 未入金 → 対象外
            ryokin = _to_money(row[COL_RYOKIN])
            if ryokin == 0:
                continue
            juku = row[COL_JUKUMEI]
            juku = juku.strip() if isinstance(juku, str) else (str(juku) if juku else "")
            if kid not in rows:
                rows[kid] = {"juku_candidates": set(), **{c: 0 for c in CATEGORIES}}
            if juku:
                rows[kid]["juku_candidates"].add(juku)
            rows[kid][cat] += ryokin
    wb.close()

    out: List[Dict] = []
    for kid, data in rows.items():
        candidates = data.pop("juku_candidates")
        # 塾名は最長の表記を採用（短縮形より正式名称を優先）
        juku_display = max(candidates, key=len) if candidates else ""
        paid = paid_map[kid]
        # 時刻成分は捨てて date-only にする（旧データの UTC ずれ等を防ぐ）
        paid_date = datetime(paid.year, paid.month, paid.day)
        rec = {
            "家族ID": kid,
            "塾名": juku_display,
            "対象月": target_month,
            "入金日": paid_date,
            **data,
            "合計": sum(data.values()),
        }
        out.append(rec)
    return out


def extract_all(send_specs: List[Dict]) -> List[Dict]:
    """複数送信分を順に抽出してフラット結合。

    send_specs: [{"path", "target_month"}, ...]
    """
    all_records = []
    for s in send_specs:
        recs = extract_sales(s["path"], s["target_month"])
        all_records.extend(recs)
    return all_records
